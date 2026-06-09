import re
import json
from openpyxl import load_workbook
from collections import defaultdict
import sys
import os

file_sche = sys.argv[1]

FILE = file_sche + ".xlsx"

wb = load_workbook(FILE, data_only=True)

results = []

dir_name = file_sche+"_data"

os.makedirs(dir_name, exist_ok=True)

# ✅ NEW: store subject names separately
subject_names = defaultdict(list)

# 🔁 Iterate through all sheets
for SHEET_NAME in wb.sheetnames:
    ws = wb[SHEET_NAME]
    row_ptr = 1  # reset pointer per sheet

    while row_ptr <= ws.max_row:
        start_row = None

        # 1️⃣ Find next "ใบรายชื่อ"
        for row in range(row_ptr, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col).value
                if isinstance(cell, str) and "ใบรายชื่อ" in cell:
                    start_row = row
                    break
            if start_row:
                break

        if not start_row:
            break  # ❌ no more "ใบรายชื่อ" in this sheet

        # 2️⃣ Find header row + metadata
        header_row = None
        room = subject = section = time = None

        for r in range(start_row, start_row + 10):
            val_a = ws[f"A{r}"].value
            val_b = ws[f"B{r}"].value

            if val_a and "รายวิชา" in str(val_a):

                # ✅ CHANGED PART (split on first space)
                subject_raw = ws[f"B{r}"].value or ""
                subject_raw = subject_raw.strip()
                subject_raw = subject_raw.replace(": ", "")

                if " " in subject_raw:
                    subject, subject_name = subject_raw.split(" ", 1)
                elif " " in subject_raw:
                    subject, subject_name = subject_raw.split(" ", 1)
                else:
                    subject = subject_raw
                    subject_name = ""

                if subject:
                    subject_names[subject].append(subject_name)

                section_raw = ws[f"F{r}"].value or ""
                section = re.sub(r"SEC\. ?", "", str(section_raw))

            if val_a and "เวลาสอบ" in str(val_a):
                time_raw = ws[f"B{r}"].value or None
                m = re.sub(" น\.", "", time_raw) if time_raw else None
                time = m

            if val_a and "ห้องสอบ" in str(val_a):
                room = ws[f"B{r}"].value

            if val_b and "รหัส" in str(val_b):
                header_row = r
                break

        if not header_row:
            row_ptr = start_row + 1
            continue

        data_row = header_row + 1

        # 3️⃣ Read student rows
        while True:
            number = ws[f"A{data_row}"].value
            student_id = ws[f"B{data_row}"].value
            branch = ws[f"D{data_row}"].value
            seat = ws[f"E{data_row}"].value
            note = ws[f"F{data_row}"].value

            if not student_id and not seat:
                break

            results.append({
                "sheet": SHEET_NAME,
                "id": int(number) if number else None,
                "time": str(time).strip() if time else None,
                "room": str(room).strip() if room else None,
                "subject": str(subject).strip() if subject else None,
                "section": str(section).strip() if section else None,
                "student_id": str(student_id).strip(),
                "branch": str(branch).strip() if branch else None,
                "seat": str(seat).strip(),
                "note": str(note).strip() if note else None,
            })

            data_row += 1

        # 4️⃣ Move pointer forward
        row_ptr = data_row # + 1
        print(row_ptr)


# ✅ Write main exam data
with open(dir_name+"/"+file_sche+".json", "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=2)

# ✅ Remove duplicate subject names per subject code (optional but safe)
subject_names = {
    k: list(dict.fromkeys(v))
    for k, v in subject_names.items()
}

# ✅ Write subject names to a separate file
with open(dir_name+"/"+file_sche+"_subjects.json", "w", encoding="utf-8") as f:
    json.dump(subject_names, f, ensure_ascii=False, indent=2)

# ===============================
# Student-based file generation
# ===============================

students_data = defaultdict(list)

# Group the data by student_id
for entry in results:
    student_id = entry['student_id']
    students_data[student_id].append(entry)

# Function to write each student's data to a separate JSON file
def write_to_file(student_id, student_data):
    filename = f"student_{student_id}.json"
    with open(file_sche+"_data/" + filename, 'w', encoding='utf-8') as f:
        json.dump(student_data, f, ensure_ascii=False, indent=4)
    print(f"Data for student {student_id} has been written to {filename}")

# Write each student's data to a separate file
for student_id, student_data in students_data.items():
    write_to_file(student_id, student_data)
